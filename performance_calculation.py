import xml.etree.ElementTree as ET
import os
from openpyxl import Workbook, load_workbook


type_CarFlow = "Auto"
type_EVFlow = "EmergencyVehicle"

#id_SideCarFlow = "SideCar"

# for conflict in SSM_root.findall('conflict'):
#     begin = conflict.attrib['begin']
#     end = conflict

EV_arrivalEdgeReferenceList = {'C4E4': ['E2N1', 'E3J2'], 'J2E3': ['E2N1', 'E1C1'],
                              'N1E2': ['E1C1', 'E4C4'], 'C1E1': ['E3J2', 'E4C4']}

enconterTypeDict = {0: 'NOCONFLICT_AHEAD', 1: '1. FOLLOWING', 2: '2. Ego車跟Foe車', 3: '3. Foe車跟Ego車', 4: '4. Foe車在隔壁車道與本車同方向行駛',
                5: '5. 匯入(A)', 6: '6. 匯入(B)', 7: '7. 匯入(C)', 8: '7. 匯入(D)', 9: '9. 橫交叉',
                10:'10. 橫交叉', 11:'11. 橫交叉', 12:'12', 13:'13', 14: '14', 15: '15',
                16: '16', 17: '17', 18: '18', 19: '19', 20: '20', 111: '111', 'NA': 'NA'}

class fileNameException(Exception):
    def __init__(self):
        self.fileName = fileName

    def __str__(self):
        return "非預期檔名: {0}".format(self.fileName)

def SSM_Run(root, outputFileName, orderOfFile):

    wb = load_workbook(output_fileName)
    ws = wb.active
    conflictOrder = 0

    def tryConvertToFloat(value):
        try:
            float(value)
            return float(value)
        except ValueError:
            return value

    for conflict in root.findall('conflict'):

        conflictOrder = conflictOrder + 1

        conflictStartTime = float(conflict.attrib['begin'])
        conflictEndTime = float(conflict.attrib['end'])
        conflict_EgoVehID = conflict.attrib['ego']
        conflict_FoeVehID = conflict.attrib['foe']

        minTTC_Time = tryConvertToFloat(conflict.find('minTTC').attrib['time'])
        minTTC_Type = enconterTypeDict[tryConvertToFloat(conflict.find('minTTC').attrib['type'])]
        #minTTC_Type = tryConvertToFloat(conflict.find('minTTC').attrib['type'])
        minTTC_Value = tryConvertToFloat(conflict.find('minTTC').attrib['value'])

        maxDRAC_Time = tryConvertToFloat(conflict.find('maxDRAC').attrib['time'])
        maxDRAC_Type = enconterTypeDict[tryConvertToFloat(conflict.find('maxDRAC').attrib['type'])]
        #maxDRAC_Type = tryConvertToFloat(conflict.find('maxDRAC').attrib['type'])
        maxDRAC_Value = tryConvertToFloat(conflict.find('maxDRAC').attrib['value'])

        # ws.append(['Conflict Order', 'Ego', 'Foe', 'Begin', 'End', 'MinTTC type',
        #                            'MinTTC time',  'MinTTC value', 'MaxDRAC Type', 'MaxDRAC Time', 'MaxDRAC Value'])
        ws.append([orderOfFile, conflictOrder, conflict_EgoVehID, conflict_FoeVehID, conflictStartTime, conflictEndTime, minTTC_Type,
                    minTTC_Time, minTTC_Value, maxDRAC_Type, maxDRAC_Time, maxDRAC_Value])
        wb.save(outputFileName)
        wb.close()

    #globalMeasure
    ws.append(['orderOfFile', 'egoVehID', 'maxBR_Value', 'maxBR_Time', 'minSGAP_Value',
               'minSGAP_Time', 'minSGAP_Leader', 'minTGAP_Value', 'minTGAP_Time', 'minTGAP_Leader'])

    for globalMeasure in root.findall('globalMeasures'):

        egoVehID = globalMeasure.attrib['ego']
        maxBR_Value = tryConvertToFloat(globalMeasure.find('maxBR').attrib['value'])
        maxBR_Time = tryConvertToFloat(globalMeasure.find('maxBR').attrib['time'])

        minSGAP_Value = tryConvertToFloat(globalMeasure.find('minSGAP').attrib['value'])
        minSGAP_Time = tryConvertToFloat(globalMeasure.find('minSGAP').attrib['time'])
        minSGAP_Leader = globalMeasure.find('minSGAP').attrib['leader']

        minTGAP_Value = tryConvertToFloat(globalMeasure.find('minTGAP').attrib['value'])
        minTGAP_Time = tryConvertToFloat(globalMeasure.find('minTGAP').attrib['time'])
        minTGAP_Leader = globalMeasure.find('minTGAP').attrib['leader']

        ws.append([orderOfFile, egoVehID, maxBR_Value, maxBR_Time, minSGAP_Value,
             minSGAP_Time, minSGAP_Leader, minTGAP_Value, minTGAP_Time, minTGAP_Leader])
        wb.save(outputFileName)
        wb.close()


def tripInfo_Run(CONTROL_STRATEGY, root, outputFileName, orderOfFile):

    EVdepartEdge = 0
    EVarrivalEdge = 0
    EVdepartTime = 0
    EVarrivalTime = 0

    samePhaseVeh_waitTime = 0
    nonSamePhaseVeh_waitTime = 0
    EV_waitTime = 0

    samePhaseVeh_WaitCount = 0
    nonSamePhaseVeh_WaitCount = 0
    EV_WaitCount = 0

    num_Of_samePhaseVeh = 0
    num_Of_nonSamePhaseVeh = 0
    num_Of_EV = 0

    def checkIsSamePhaseVehicle(vehInfo, EVInfo, controlStrategy):
        vehDepartTime = vehInfo[0]
        vehArrivalTime = vehInfo[1]
        vehDepartEdge = vehInfo[2]
        vehArrivalEdge = vehInfo[3]

        EVdepartTime = EVInfo[0]
        EVarrivalTime = EVInfo[1]
        EVdepartEdge = EVInfo[2]
        EVarrivalEdge = EVInfo[3]

        if controlStrategy in [0, 1]:  # 綠燈延長 紅燈縮短
            if vehDepartTime >= (EVdepartTime - 30) and vehDepartTime <= (EVarrivalTime + 30)  \
                    and vehDepartEdge == EVdepartEdge and vehArrivalEdge in EV_arrivalEdgeReferenceList[vehDepartEdge]:
                print("controlStrategy = ", controlStrategy, "與EV同向車")
                return True
            else: return False
        elif controlStrategy in [2, 3]:  # 插入時相
            if vehDepartTime >= (EVdepartTime - 30) and vehDepartTime <= (EVarrivalTime + 30) and vehDepartEdge == EVdepartEdge:
                print("controlStrategy = ", controlStrategy, "與EV同向車")
                return True
            else: return False
        else:
            print(1/0)

    for tripinfo in root.findall('tripinfo'):

        vehID = tripinfo.get('id')
        vehDepartTime = float(tripinfo.get('depart'))
        vehArrivalTime = float(tripinfo.get('arrival'))
        vehDepartEdge = tripinfo.get('departLane')[:4]
        vehArrivalEdge = tripinfo.get('arrivalLane')[:4]

        if vehArrivalTime >= 7100:
            continue

        if tripinfo.get('vType') == 'EmergencyVehicle':
            EV_id = tripinfo.get('id')
            EVdepartTime = float(tripinfo.get('depart'))
            EVarrivalTime = float(tripinfo.get('arrival'))
            EVdepartEdge = tripinfo.get('departLane')[:4]
            EVarrivalEdge = tripinfo.get('arrivalLane')[:4]
            EV_waitTime = EV_waitTime + float(tripinfo.get('waitingTime'))
            EV_WaitCount = EV_WaitCount + int(tripinfo.get('waitingCount'))
            num_Of_EV = num_Of_EV + 1
            continue

        vehInfo = [vehDepartTime, vehArrivalTime, vehDepartEdge, vehArrivalEdge]
        EVInfo = [EVdepartTime, EVarrivalTime, EVdepartEdge, EVarrivalEdge]
        IsSamePhaseVehicle = checkIsSamePhaseVehicle(vehInfo=vehInfo, EVInfo=EVInfo, controlStrategy=CONTROL_STRATEGY)
        vehWaitingTime = float(tripinfo.get('waitingTime'))
        vehStopCount = int(tripinfo.get('waitingCount'))

        if IsSamePhaseVehicle:
            samePhaseVeh_waitTime = samePhaseVeh_waitTime + vehWaitingTime
            samePhaseVeh_WaitCount = samePhaseVeh_WaitCount + vehStopCount
            num_Of_samePhaseVeh = num_Of_samePhaseVeh + 1
        else:
            nonSamePhaseVeh_waitTime = nonSamePhaseVeh_waitTime + vehWaitingTime
            nonSamePhaseVeh_WaitCount = nonSamePhaseVeh_WaitCount + vehStopCount
            num_Of_nonSamePhaseVeh = num_Of_nonSamePhaseVeh + 1

    print("------------------------------")
    try:
        samePhaseVeh_delay = round((samePhaseVeh_waitTime / num_Of_samePhaseVeh), 2)
        nonSamePhaseVeh_delay = round((nonSamePhaseVeh_waitTime / num_Of_nonSamePhaseVeh), 2)
        EV_delay = round((EV_waitTime / num_Of_EV), 2)

        samePhase_Avg_WaitCount = round((samePhaseVeh_WaitCount/num_Of_samePhaseVeh), 2)
        nonSamePhase_Avg_WaitCount = round((nonSamePhaseVeh_WaitCount / num_Of_nonSamePhaseVeh), 2)
        EV_Avg_waitCount = round((EV_WaitCount / num_Of_EV), 2)

        totalWaitTime = nonSamePhaseVeh_waitTime + samePhaseVeh_waitTime + EV_waitTime
        totalWaitCount = samePhaseVeh_WaitCount + nonSamePhaseVeh_WaitCount + EV_WaitCount
        totalAmountOfCar = num_Of_EV + num_Of_samePhaseVeh + num_Of_nonSamePhaseVeh

        total_avg_delay = round((totalWaitTime / totalAmountOfCar), 2)
        total_avg_waitCount = round((totalWaitCount / totalAmountOfCar), 2)

        print("------Summary-----")
        print("SamePhaseVeh delay=", samePhaseVeh_delay)
        print("nonSamePhaseVeh delay=", nonSamePhaseVeh_delay)
        print("EV delay=", EV_delay)
        print("SamePhaseVeh WaitCount = ", samePhase_Avg_WaitCount)
        print("nonSamePhaseVeh WaitCount = ", nonSamePhase_Avg_WaitCount)
        print("EV WaitCount = ", EV_Avg_waitCount)
        print("Total avg delay=", total_avg_delay)
        print("Total avg WaitCount=", total_avg_waitCount)

    except:
        raise ZeroDivisionError("!!!")

    with open(outputFileName, 'a+', newline='') as csvfile:
        # 建立 CSV 檔寫入器
        writer = csv.writer(csvfile)
        # 寫入資料
        writer.writerow([orderOfFile, 'Delay', EV_delay, samePhaseVeh_delay, nonSamePhaseVeh_delay, total_avg_delay])
        #writer.writerow([orderOfFile, 'WaitCount', EV_Avg_waitCount, samePhase_Avg_WaitCount, nonSamePhase_Avg_WaitCount, total_avg_waitCount])

if __name__ == "__main__":

    def inputSelecter():

        VC = input("V/C = 1. 0.9 2. 0.5\n")
        if VC == '1': VC = 'VC_0.9'
        else: VC = 'VC_0.5'

        EV_level = input("EV level -> 1 = High 2 = Medium 3 = Low\n")
        if EV_level == '1': EV_level = 'EV_High'
        elif EV_level == '2': EV_level = 'EV_Medium'
        elif EV_level == '3': EV_level = 'EV_Low'

        LeftLevel = input("Left Level -> 1 = High 2 = Medium 3 = Low\n")
        if LeftLevel == '1': LeftLevel = 'HighLeft'
        elif LeftLevel == '2': LeftLevel = 'MediumLeft'
        elif LeftLevel == '3': LeftLevel = 'LowLeft'

        CONTROL_STRATEGY = int(input("ControlStrategy = "))

        inputParameters = {'VC': VC, 'EV_level': EV_level, 'LeftLevel': LeftLevel, 'ControlStrategy': CONTROL_STRATEGY}

        return inputParameters

    inputParameters = inputSelecter()

    # VC = ['1', '3']
    # EV_LEVEL = ['1', '2', '3']
    # LEFT_LEVEL = ['1', '3']
    # CONTROL_STRATEGY_List = ['1', '2', '3']
    #
    # for vc in VC:
    #     for EV_level in EV_LEVEL:
    #         for LeftLevel in LEFT_LEVEL:
    #             for CONTROL_STRATEGY in CONTROL_STRATEGY_List:
    #

    route = inputParameters['VC'] + '/' + inputParameters['EV_level'] + '/' + inputParameters['LeftLevel'] + '/' + 'result' + '/' + str(inputParameters['ControlStrategy']) + '/'

    from os import listdir
    tripInfoFileName = 0
    SSMFileName = 0
    old_fileName = 0
    for fileName in listdir(path=route): # 取得目錄底下所有檔案名稱
        if fileName[-12:-4] == 'tripInfo':  # 解析檔案名稱
            tripInfoFileName = fileName
            tripInfo_tree = ET.ElementTree(file=route + tripInfoFileName)
            tripInfo_root = tripInfo_tree.getroot()
            import csv

            routeNew = route.replace('/', '_')

            # 開啟輸出的 excel 檔案
            output_fileName = "output/" + "traffic_Output" + "_" + routeNew + ".xlsx"
            if output_fileName != old_fileName:
                wb = Workbook()
                ws = wb.active
                ws.append(['Order', 'Type', 'EV', 'SamePhaseVeh', 'NonSamePhaseVeh', 'Overall'])
                wb.save(output_fileName)
                wb.close()
                old_fileName = output_fileName

            #tripInfo_Run(CONTROL_STRATEGY=inputParameters['ControlStrategy'], root=tripInfo_root, outputFileName=output_fileName,orderOfFile=fileName[:3])
            continue

        elif fileName[-7:-4] == 'SSM':
            SSMFileName = fileName
            SSM_tree = ET.ElementTree(file=route + SSMFileName)
            SSM_root = SSM_tree.getroot()

            routeNew = route.replace('/', '_')

            output_fileName = "output/" + "SSM_Output" + "_" + routeNew + ".xlsx"

            if output_fileName != old_fileName:

                wb = Workbook()
                ws = wb.active
                ws.append(['File', 'Conflict Order', 'Ego', 'Foe', 'Begin', 'End', 'MinTTC type',
                           'MinTTC time',  'MinTTC value', 'MaxDRAC Type', 'MaxDRAC Time', 'MaxDRAC Value'])
                #ws.column_dimensions[0].width = 20.0
                wb.save(output_fileName)
                wb.close()
                old_fileName = output_fileName

            SSM_Run(root=SSM_root, outputFileName=output_fileName, orderOfFile=fileName[:3])
            continue

        else: #有非預期檔案在資料夾中 -> 報錯
            raise fileNameException

    def converCSVtoEXCEL(fileRoute):
        import pandas as pd
        read_file = pd.read_csv(fileRoute, encoding='utf-8')
        read_file.to_excel(fileRoute + '.xlsx', index=None, header=True)

    #$converCSVtoEXCEL(fileRoute=old_fileName)





